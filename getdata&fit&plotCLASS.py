#import random
import math
#import pandas
#import requests
import openpyxl
import matplotlib.pyplot as plt
import numpy as np

def sigma_sum(start: int, end: int, expression) -> float:
    return sum(expression(i) for i in range(start, end + 1))

class DataFitterPlotter:
    def __init__(self, workbook: str, worksheet: str) -> None:
        # Set the font to Times New Roman
        plt.rcParams["font.family"] = "Times New Roman"

        #Doesn't paste data into spreadsheet, just prints them out
        #doesn't create new xlxs file
        #workbook = xlsxwriter.Workbook("PI-LAB2025.xlxs")
        self.workbook = openpyxl.load_workbook(workbook)

        #there is some bug when trying to use Sheet 2 for y=ax, P=0...
        self.worksheet = self.workbook[worksheet]
        self.get_data()

        self.modified_delta_y_values = []

        self.initial_a_yax = 0
        self.final_a_yax = 0
        self.delta_a_yax = 0

        self.initial_a_yaxplusb = 0
        self.final_a_yaxplusb = 0
        self.delta_a_yaxplusb = 0
        self.b = 0
        self.delta_b = 0

        self.A = 1
        self.delta_A = 0

        self.number_of_parameters = 1
        self.degrees_of_freedom = 1
        self.chi_squared = 0
        self.reduced_chi_squared = 0
                
    def get_data(self) -> None:
        self.x_values = []
        self.y_values = []
        self.delta_y_values = []
        self.delta_x_values = []

        self.x_parameter = self.worksheet['A1'].value
        self.y_parameter = self.worksheet["B1"].value
        self.delta_x_parameter = self.worksheet["C1"].value
        self.delta_y_parameter = self.worksheet["D1"].value

        for value in self.worksheet.iter_rows(min_row=2,
                                        min_col=1,
                                        max_col=1,
                                        values_only=True):
            if value[0] != None:  
                self.x_values.append(value[0])
            else: break

        for value in self.worksheet.iter_rows(min_row=2,
                                        min_col=2,
                                        max_col=2,
                                        values_only=True):
            if value[0] != None:
                self.y_values.append(value[0])
            else: break

        for value in self.worksheet.iter_rows(min_row=2,
                                        min_col=4,
                                        max_col=4,
                                        values_only=True):
            if value[0] != None:
                self.delta_y_values.append(value[0])
            else: break

        for value in self.worksheet.iter_rows(min_row=2,
                        max_row=100,
                        min_col=3,
                        max_col=3,
                        values_only=True):
            if value[0] != None:
                self.delta_x_values.append(value[0])
            else: break
  
    def calculate_parameters_yax(self, p: float=0, manual_delta_y: float=0.1) -> None:
        float(p)
        float(manual_delta_y)
        
        print("FITTING y = ax:")
        self.b = 0
        self.number_of_parameters = 1
        self.degrees_of_freedom = len(self.x_values) - self.number_of_parameters

        if p!=0:
            #Manually try better delta_y_values to get reduced_chi_squared between 0-1 (depending of degrees of freedom)
            for i in range(len(self.modified_delta_y_values)):
                #make manual entered number
                self.modified_delta_y_values[i] = manual_delta_y
            print(f"\ndelta_y_values have been revised (new delta_y's={self.modified_delta_y_values[0]})...\n")
            print(self.modified_delta_y_values)

            print(f"REFITTING y=ax for {p=}...")
            self.final_a_yax =  sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]*self.y_values[i])/(self.modified_delta_y_values[i]**2)) / sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]/self.modified_delta_y_values[i])**2)
            self.delta_a_yax =  1/math.sqrt(sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]/self.modified_delta_y_values[i])**2))

            self.chi_squared = sigma_sum(0, len(self.x_values)-1, lambda i: ((1/(self.modified_delta_y_values[i]))*((self.y_values[i] - self.final_a_yax * self.x_values[i] - self.b)))**2)
            self.reduced_chi_squared = self.chi_squared / self.degrees_of_freedom
            print(f"\nRecalculated delta y's, a, b, delta a, delta, b, chi sqaured, reduced chi-squared:")
            print(f"{p=}, {self.modified_delta_y_values=}, \n {self.final_a_yax=}, {self.b=}, {self.delta_a_yax=}, \n {self.chi_squared=}, {self.reduced_chi_squared=}")
    
        else:
            self.initial_a_yax =  sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]*self.y_values[i])/(self.delta_y_values[i]**2)) / sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]/self.delta_y_values[i])**2)
            self.delta_a_yax =  1/math.sqrt(sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]/self.delta_y_values[i])**2))
            
            if len(self.modified_delta_y_values)==0:
                for i in range(len(self.delta_y_values)):
                    self.modified_delta_y_values.append(math.sqrt((self.delta_y_values[i])**2 + (self.initial_a_yax*self.delta_x_values[i])**2))
                    self.modified_delta_y_values[i] =  self.modified_delta_y_values[i]
            else:
                for i in range(len(self.delta_y_values)):
                    self.modified_delta_y_values[i] = math.sqrt((self.delta_y_values[i])**2 + (self.initial_a_yax*self.delta_x_values[i])**2)
                    self.modified_delta_y_values[i] =  self.modified_delta_y_values[i]

            self.final_a_yax =  sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]*self.y_values[i])/(self.modified_delta_y_values[i]**2)) / sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]/self.modified_delta_y_values[i])**2)
            self.delta_a_yax =  1/math.sqrt(sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]/self.modified_delta_y_values[i])**2))

            self.chi_squared = sigma_sum(0, len(self.x_values)-1, lambda i: ((1/(self.modified_delta_y_values[i]))*((self.y_values[i] - self.final_a_yax * self.x_values[i] - self.b)))**2)
            self.reduced_chi_squared = self.chi_squared / self.degrees_of_freedom

            #DEBUG
            print(f"{self.initial_a_yax=}, {self.final_a_yax}, {self.delta_a_yax=}, {self.chi_squared=}, \n {self.reduced_chi_squared=},  {self.delta_y_parameter} :: {self.delta_y_values=}, \n {self.delta_x_parameter} :: {self.delta_x_values=}")
            print(f"{self.x_parameter} :: {self.x_values=}, {self.y_parameter} :: {self.y_values=},\n{self.modified_delta_y_values=}, \n{self.degrees_of_freedom=}, {self.final_a_yax=}, {self.delta_a_yax=}, \n {self.chi_squared=}, {self.reduced_chi_squared=}")

    def calculate_parameters_yaxplusb(self, p: float=0, manual_delta_y: float=0.1) -> None:
        float(p)
        float(manual_delta_y)
        
        #y=ax+b
        self.number_of_parameters = 2
        self.degrees_of_freedom = len(self.x_values) - self.number_of_parameters

        if p!=0:
            #Manually try better delta_y_values to get reduced_chi_squared between 0-1 (depending of degrees of freedom)
            for i in range(len(self.modified_delta_y_values)):
                #make manual entered number
                self.modified_delta_y_values[i] = manual_delta_y
            print(f"REFITTING y=ax+b for {p=}...")
            self.triangle = sigma_sum(0, len(self.x_values)-1, lambda i: (1/self.modified_delta_y_values[i])**2)*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]/self.modified_delta_y_values[i])**2) - (sigma_sum(0, len(self.x_values)-1, lambda i: self.x_values[i]/(self.modified_delta_y_values[i]**2)))**2
            self.final_a_yaxplusb = (1/self.triangle)*(sigma_sum(0, len(self.x_values)-1, lambda i: 1/(self.modified_delta_y_values[i]**2))*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]*self.y_values[i])/(self.modified_delta_y_values[i]**2)) - sigma_sum(0, len(self.x_values)-1, lambda i: self.y_values[i]/(self.modified_delta_y_values[i]**2))*sigma_sum(0, len(self.x_values)-1, lambda i: self.x_values[i]/(self.modified_delta_y_values[i]**2)))

            self.b = (1/self.triangle)*((sigma_sum(0,len(self.x_values)-1, lambda i: self.y_values[i]/(self.modified_delta_y_values[i]**2))*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]/self.modified_delta_y_values[i])**2)) - (sigma_sum(0, len(self.x_values)-1, lambda i: self.x_values[i]/(self.modified_delta_y_values[i]**2))*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]*self.y_values[i])/(self.modified_delta_y_values[i]**2))))
            self.delta_a_yaxplusb = math.sqrt((1/self.triangle)*sigma_sum(0, len(self.x_values)-1, lambda i: (1/(self.modified_delta_y_values[i]**2))))
            self.delta_b = math.sqrt((1/self.triangle)*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]/self.modified_delta_y_values[i])**2))

            self.chi_squared = sigma_sum(0, len(self.x_values)-1, lambda i: ((1/(self.modified_delta_y_values[i]))*((self.y_values[i] - self.final_a_yaxplusb * self.x_values[i] -self.b)))**2)
            self.reduced_chi_squared = self.chi_squared / self.degrees_of_freedom
            print(f"\nRecalculated delta y's, a, b, delta a, delta, b, chi sqaured, reduced chi-squared:")
            print(f"{p=}, {self.modified_delta_y_values=}, \n {self.final_a_yaxplusb=}, {self.b=}, {self.delta_a_yaxplusb=}, {self.delta_b=}, \n {self.chi_squared=}, {self.reduced_chi_squared=}")

        else:
            print("\n\n FITTING y = ax + b (initial):\n")
            self.triangle = sigma_sum(0, len(self.x_values)-1, lambda i: (1/self.delta_y_values[i])**2)*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]/self.delta_y_values[i])**2) - (sigma_sum(0, len(self.x_values)-1, lambda i: self.x_values[i]/(self.delta_y_values[i]**2)))**2
            self.t1 = sigma_sum(0, len(self.x_values)-1, lambda i: (1/self.delta_y_values[i])**2)*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]/self.delta_y_values[i])**2)
            self.t2 = (sigma_sum(0, len(self.x_values)-1, lambda i: self.x_values[i]/(self.delta_y_values[i]**2)))**2
            
            print(f"{self.t1=}, {self.t2=}, {self.triangle=}, {len(self.x_values)}, {self.number_of_parameters=}, {self.degrees_of_freedom=}")

            self.initial_a_yaxplusb = (1/self.triangle)*(sigma_sum(0, len(self.x_values)-1, lambda i: 1/(self.delta_y_values[i]**2))*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]*self.y_values[i])/(self.delta_y_values[i]**2)) - sigma_sum(0, len(self.x_values)-1, lambda i: self.y_values[i]/(self.delta_y_values[i]**2))*sigma_sum(0, len(self.x_values)-1, lambda i: self.x_values[i]/(self.delta_y_values[i]**2)))
            self.b = (1/self.triangle)*((sigma_sum(0,len(self.x_values)-1, lambda i: self.y_values[i]/(self.delta_y_values[i]**2))*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]/self.delta_y_values[i])**2)) - (sigma_sum(0, len(self.x_values)-1, lambda i: self.x_values[i]/(self.delta_y_values[i]**2))*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]*self.y_values[i])/(self.delta_y_values[i]**2))))
            self.delta_a_yaxplusb = math.sqrt((1/self.triangle)*sigma_sum(0, len(self.x_values)-1, lambda i: (1/(self.delta_y_values[i]**2))))
            self.delta_b = math.sqrt((1/self.triangle)*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]/self.delta_y_values[i])**2))

            self.chi_squared = sigma_sum(0, len(self.x_values)-1, lambda i: ((1/(self.delta_y_values[i]))*((self.y_values[i] - self.initial_a_yaxplusb * self.x_values[i] - self.b)))**2)
            self.reduced_chi_squared = self.chi_squared / self.degrees_of_freedom
            
            # WRUTOUT INIT
            print(f"{self.x_values=}, {self.y_values=}, {self.initial_a_yaxplusb=}, \n {self.b=}, \n {self.delta_a_yaxplusb=}, {self.delta_b=}")
            print(f"{self.delta_y_values=}, {self.modified_delta_y_values=}, {self.degrees_of_freedom=}, {self.chi_squared=}, {self.reduced_chi_squared=}")

            #final - new deltas, new a, new b
            if len(self.modified_delta_y_values)==0:
                for i in range(len(self.delta_y_values)):
                    self.modified_delta_y_values.append(math.sqrt((self.delta_y_values[i])**2 + (self.initial_a_yaxplusb*self.delta_x_values[i])**2))
                    self.modified_delta_y_values[i] =  self.modified_delta_y_values[i]
                for i in range(len(self.delta_y_values)):
                    self.modified_delta_y_values[i] = math.sqrt((self.delta_y_values[i])**2 + (self.initial_a_yaxplusb*self.delta_x_values[i])**2)
                    self.modified_delta_y_values[i] =  self.modified_delta_y_values[i]

            print("\n\n FITTING y = ax + b (final):\n")

            self.triangle = sigma_sum(0, len(self.x_values)-1, lambda i: (1/self.modified_delta_y_values[i])**2)*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]/self.modified_delta_y_values[i])**2) - (sigma_sum(0, len(self.x_values)-1, lambda i: self.x_values[i]/(self.modified_delta_y_values[i]**2)))**2
            self.final_a_yaxplusb = (1/self.triangle)*(sigma_sum(0, len(self.x_values)-1, lambda i: 1/(self.modified_delta_y_values[i]**2))*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]*self.y_values[i])/(self.modified_delta_y_values[i]**2)) - sigma_sum(0, len(self.x_values)-1, lambda i: self.y_values[i]/(self.modified_delta_y_values[i]**2))*sigma_sum(0, len(self.x_values)-1, lambda i: self.x_values[i]/(self.modified_delta_y_values[i]**2)))
            
            self.b = (1/self.triangle)*((sigma_sum(0,len(self.x_values)-1, lambda i: self.y_values[i]/(self.modified_delta_y_values[i]**2))*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]/self.modified_delta_y_values[i])**2)) - (sigma_sum(0, len(self.x_values)-1, lambda i: self.x_values[i]/(self.modified_delta_y_values[i]**2))*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]*self.y_values[i])/(self.modified_delta_y_values[i]**2))))
            self.delta_a_yaxplusb = math.sqrt((1/self.triangle)*sigma_sum(0, len(self.x_values)-1, lambda i: (1/(self.modified_delta_y_values[i]**2))))
            self.delta_b = math.sqrt((1/self.triangle)*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]/self.modified_delta_y_values[i])**2))

            self.chi_squared = sigma_sum(0, len(self.x_values)-1, lambda i: ((1/(self.modified_delta_y_values[i]))*((self.y_values[i] - self.final_a_yaxplusb * self.x_values[i] - self.b)))**2)
            self.reduced_chi_squared = self.chi_squared / self.degrees_of_freedom

        #The formula for calculating P for given reduced-chi-squared is more effort than I want: See RB Eq. 11.6

        # WRITOUT FINAL
        print(f"{self.modified_delta_y_values=}, \n {self.final_a_yaxplusb=}, {self.b=}, {self.delta_a_yaxplusb=}, {self.delta_b=}, \n {self.chi_squared=}, {self.reduced_chi_squared=}")

    def plot_y_ax(self, p: float=0) -> None:
        float(p)

        if (p == 0):
            plt.figure(1)
        else:
            plt.figure(3)

        # Fit line
        x_line = np.linspace(min(self.x_values)-abs(min(self.x_values))*0.2, max(self.x_values) * 1.1, 200)
        y_line = self.final_a_yax * x_line + self.b

        # Plot with hollow markers and whiskers
        plt.errorbar(
            self.x_values, self.y_values,
            yerr=self.modified_delta_y_values,
            fmt='o', markersize=4,
            mfc='none', mec='black',
            ecolor='black', elinewidth=1.5, capsize=2,
            label='Data with uncertainties'
        )

        # Fit line
        plt.plot(x_line, y_line, color='blue', linewidth=2, label=f'Fit: y = {self.final_a_yax:.4f}x')

        # Labels
        plt.xlabel(f"{self.x_parameter}")
        plt.ylabel(f"{self.y_parameter}")

        if (p!=0):
            plt.title(f"Linear Least-Squares Fit of Measurements (y = ax) ensuring P = {p}")
        else:
            plt.title(f"Linear Least-Squares Fit of Measurements (y = ax)")

        # Add chi-squared summary box
        textstr = '\n'.join((
            r'$\chi^2 = %.3f$' % (self.chi_squared,),
            r'$\chi^2_\nu = %.3f$' % (self.reduced_chi_squared,)
        ))

        plt.gca().text(
            0.05, 0.95, textstr,
            transform=plt.gca().transAxes,
            fontsize=10,
            verticalalignment='top',
            bbox=dict(boxstyle='round,pad=0.3', facecolor='white', alpha=0.7)
        )

        plt.legend()
        plt.grid(True)
        #plt.show()

    def plot_y_axplusb(self, p: float=0) -> None: 
        float(p)

        if (p == 0):
            plt.figure(2)
        else:
            plt.figure(4)

        # Fit line
        x_line = np.linspace(min(self.x_values)-abs(min(self.x_values))*0.2, max(self.x_values) * 1.1, 200)
        y_line = self.final_a_yaxplusb * x_line + self.b

        # Plot with error bars
        plt.errorbar(
            self.x_values, self.y_values, 
            yerr=self.modified_delta_y_values,
            fmt='o', markersize=4,
            mfc='none', mec='black',
            ecolor='black', elinewidth=1.5, capsize=2,
            label='Data with uncertainties'
        )

        # Fit line
        plt.plot(x_line, y_line, color='green', linewidth=2, label=f'Fit: y = {self.final_a_yaxplusb:.2f}x + {self.b:.2f}')

        # Labels
        plt.xlabel(f"{self.x_parameter}")
        plt.ylabel(f"{self.y_parameter}")

        if (p!=0):
            plt.title(f"Linear Least-Squares Fit of Measurements (y=ax+b) ensuring P = {p}")
        else:
            plt.title("Linear Least-Squares Fit of Measurements (y=ax+b)")

        # Add chi-squared summary box
        textstr = '\n'.join((
            r'$\chi^2 = %.4f$' % (self.chi_squared,),
            r'$\chi^2_\nu = %.4f$' % (self.reduced_chi_squared,),
        ))
        plt.gca().text(
            0.05, 0.50, textstr,
            transform=plt.gca().transAxes,
            fontsize=10,
            verticalalignment='top',
            bbox=dict(boxstyle='round,pad=0.3', facecolor='white', alpha=0.7)
        )

        plt.legend()
        plt.grid(True)
        #plt.show()        

#DO AS SOON AS POSSIBLE!!!!
    def calculate_parameters_yaxSQUAREDplusb(self, p: float=0, manual_delta_y: float=0.1) -> None:
        float(p)
        float(manual_delta_y)
        
        #y=ax+b
        self.number_of_parameters = 2
        self.degrees_of_freedom = len(self.x_values) - self.number_of_parameters

        if p!=0:
            #Manually try better delta_y_values to get reduced_chi_squared between 0-1 (depending of degrees of freedom)
            for i in range(len(self.modified_delta_y_values)):
                #make manual entered number
                self.modified_delta_y_values[i] = manual_delta_y
            print(f"REFITTING y=ax+b for {p=}...")
            self.triangle = sigma_sum(0, len(self.x_values)-1, lambda i: (1/self.modified_delta_y_values[i])**2)*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]/self.modified_delta_y_values[i])**2) - (sigma_sum(0, len(self.x_values)-1, lambda i: self.x_values[i]/(self.modified_delta_y_values[i]**2)))**2
            self.final_a_yaxplusb = (1/self.triangle)*(sigma_sum(0, len(self.x_values)-1, lambda i: 1/(self.modified_delta_y_values[i]**2))*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]*self.y_values[i])/(self.modified_delta_y_values[i]**2)) - sigma_sum(0, len(self.x_values)-1, lambda i: self.y_values[i]/(self.modified_delta_y_values[i]**2))*sigma_sum(0, len(self.x_values)-1, lambda i: self.x_values[i]/(self.modified_delta_y_values[i]**2)))

            self.b = (1/self.triangle)*((sigma_sum(0,len(self.x_values)-1, lambda i: self.y_values[i]/(self.modified_delta_y_values[i]**2))*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]/self.modified_delta_y_values[i])**2)) - (sigma_sum(0, len(self.x_values)-1, lambda i: self.x_values[i]/(self.modified_delta_y_values[i]**2))*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]*self.y_values[i])/(self.modified_delta_y_values[i]**2))))
            self.delta_a_yaxplusb = math.sqrt((1/self.triangle)*sigma_sum(0, len(self.x_values)-1, lambda i: (1/(self.modified_delta_y_values[i]**2))))
            self.delta_b = math.sqrt((1/self.triangle)*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]/self.modified_delta_y_values[i])**2))

            self.chi_squared = sigma_sum(0, len(self.x_values)-1, lambda i: ((1/(self.modified_delta_y_values[i]))*((self.y_values[i] - self.final_a_yaxplusb * self.x_values[i] -self.b)))**2)
            self.reduced_chi_squared = self.chi_squared / self.degrees_of_freedom
            print(f"\nRecalculated delta y's, a, b, delta a, delta, b, chi sqaured, reduced chi-squared:")
            print(f"{p=}, {self.modified_delta_y_values=}, \n {self.final_a_yaxplusb=}, {self.b=}, {self.delta_a_yaxplusb=}, {self.delta_b=}, \n {self.chi_squared=}, {self.reduced_chi_squared=}")

        else:
            print("\n\n FITTING y = ax + b (initial):\n")
            self.triangle = sigma_sum(0, len(self.x_values)-1, lambda i: (1/self.delta_y_values[i])**2)*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]/self.delta_y_values[i])**2) - (sigma_sum(0, len(self.x_values)-1, lambda i: self.x_values[i]/(self.delta_y_values[i]**2)))**2
            self.t1 = sigma_sum(0, len(self.x_values)-1, lambda i: (1/self.delta_y_values[i])**2)*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]/self.delta_y_values[i])**2)
            self.t2 = (sigma_sum(0, len(self.x_values)-1, lambda i: self.x_values[i]/(self.delta_y_values[i]**2)))**2
            
            print(f"{self.t1=}, {self.t2=}, {self.triangle=}, {len(self.x_values)}, {self.number_of_parameters=}, {self.degrees_of_freedom=}")

            self.initial_a_yaxplusb = (1/self.triangle)*(sigma_sum(0, len(self.x_values)-1, lambda i: 1/(self.delta_y_values[i]**2))*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]*self.y_values[i])/(self.delta_y_values[i]**2)) - sigma_sum(0, len(self.x_values)-1, lambda i: self.y_values[i]/(self.delta_y_values[i]**2))*sigma_sum(0, len(self.x_values)-1, lambda i: self.x_values[i]/(self.delta_y_values[i]**2)))
            self.b = (1/self.triangle)*((sigma_sum(0,len(self.x_values)-1, lambda i: self.y_values[i]/(self.delta_y_values[i]**2))*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]/self.delta_y_values[i])**2)) - (sigma_sum(0, len(self.x_values)-1, lambda i: self.x_values[i]/(self.delta_y_values[i]**2))*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]*self.y_values[i])/(self.delta_y_values[i]**2))))
            self.delta_a_yaxplusb = math.sqrt((1/self.triangle)*sigma_sum(0, len(self.x_values)-1, lambda i: (1/(self.delta_y_values[i]**2))))
            self.delta_b = math.sqrt((1/self.triangle)*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]/self.delta_y_values[i])**2))

            self.chi_squared = sigma_sum(0, len(self.x_values)-1, lambda i: ((1/(self.delta_y_values[i]))*((self.y_values[i] - self.initial_a_yaxplusb * self.x_values[i] - self.b)))**2)
            self.reduced_chi_squared = self.chi_squared / self.degrees_of_freedom
            
            # WRUTOUT INIT
            print(f"{self.x_values=}, {self.y_values=}, {self.initial_a_yaxplusb=}, \n {self.b=}, \n {self.delta_a_yaxplusb=}, {self.delta_b=}")
            print(f"{self.delta_y_values=}, {self.modified_delta_y_values=}, {self.degrees_of_freedom=}, {self.chi_squared=}, {self.reduced_chi_squared=}")

            #final - new deltas, new a, new b
            if len(self.modified_delta_y_values)==0:
                for i in range(len(self.delta_y_values)):
                    self.modified_delta_y_values.append(math.sqrt((self.delta_y_values[i])**2 + (self.initial_a_yaxplusb*self.delta_x_values[i])**2))
                    self.modified_delta_y_values[i] =  self.modified_delta_y_values[i]
            else:
                for i in range(len(self.delta_y_values)):
                    self.modified_delta_y_values[i] = math.sqrt((self.delta_y_values[i]))

            print("\n\n FITTING y = ax + b (final):\n")

            self.triangle = sigma_sum(0, len(self.x_values)-1, lambda i: (1/self.modified_delta_y_values[i])**2)*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]/self.modified_delta_y_values[i])**2) - (sigma_sum(0, len(self.x_values)-1, lambda i: self.x_values[i]/(self.modified_delta_y_values[i]**2)))**2
            self.final_a_yaxplusb = (1/self.triangle)*(sigma_sum(0, len(self.x_values)-1, lambda i: 1/(self.modified_delta_y_values[i]**2))*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]*self.y_values[i])/(self.modified_delta_y_values[i]**2)) - sigma_sum(0, len(self.x_values)-1, lambda i: self.y_values[i]/(self.modified_delta_y_values[i]**2))*sigma_sum(0, len(self.x_values)-1, lambda i: self.x_values[i]/(self.modified_delta_y_values[i]**2)))
            
            self.b = (1/self.triangle)*((sigma_sum(0,len(self.x_values)-1, lambda i: self.y_values[i]/(self.modified_delta_y_values[i]**2))*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]/self.modified_delta_y_values[i])**2)) - (sigma_sum(0, len(self.x_values)-1, lambda i: self.x_values[i]/(self.modified_delta_y_values[i]**2))*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]*self.y_values[i])/(self.modified_delta_y_values[i]**2))))
            self.delta_a_yaxplusb = math.sqrt((1/self.triangle)*sigma_sum(0, len(self.x_values)-1, lambda i: (1/(self.modified_delta_y_values[i]**2))))
            self.delta_b = math.sqrt((1/self.triangle)*sigma_sum(0, len(self.x_values)-1, lambda i: (self.x_values[i]/self.modified_delta_y_values[i])**2))

            self.chi_squared = sigma_sum(0, len(self.x_values)-1, lambda i: ((1/(self.modified_delta_y_values[i]))*((self.y_values[i] - self.final_a_yaxplusb * self.x_values[i] - self.b)))**2)
            self.reduced_chi_squared = self.chi_squared / self.degrees_of_freedom

        #The formula for calculating P for given reduced-chi-squared is more effort than I want: See RB Eq. 11.6

        # WRITOUT FINAL
        print(f"{self.modified_delta_y_values=}, \n {self.final_a_yaxplusb=}, {self.b=}, {self.delta_a_yaxplusb=}, {self.delta_b=}, \n {self.chi_squared=}, {self.reduced_chi_squared=}")

    def plot_y_axSQURAEDSSSDSDDplusb(self, p: float=0) -> None: 
        float(p)

        if (p == 0):
            plt.figure(2)
        else:
            plt.figure(4)

        # Fit line
        x_line = np.linspace(min(self.x_values)-abs(min(self.x_values))*0.2, max(self.x_values) * 1.1, 200)
        y_line = self.final_a_yaxplusb * x_line + self.b

        # Plot with error bars
        plt.errorbar(
            self.x_values, self.y_values, 
            yerr=self.modified_delta_y_values,
            fmt='o', markersize=4,
            mfc='none', mec='black',
            ecolor='black', elinewidth=1.5, capsize=2,
            label='Data with uncertainties'
        )

        # Fit line
        plt.plot(x_line, y_line, color='green', linewidth=2, label=f'Fit: y = {self.final_a_yaxplusb:.2f}x + {self.b:.2f}')

        # Labels
        plt.xlabel(f"{self.x_parameter}")
        plt.ylabel(f"{self.y_parameter}")

        if (p!=0):
            plt.title(f"Linear Least-Squares Fit of Diverging Lens (y=ax+b) ensuring P = {p}")
        else:
            plt.title("Linear Least-Squares Fit of Diverging Lens (y=ax+b)")

        # Add chi-squared summary box
        textstr = '\n'.join((
            r'$\chi^2 = %.4f$' % (self.chi_squared,),
            r'$\chi^2_\nu = %.4f$' % (self.reduced_chi_squared,),
        ))
        plt.gca().text(
            0.05, 0.50, textstr,
            transform=plt.gca().transAxes,
            fontsize=10,
            verticalalignment='top',
            bbox=dict(boxstyle='round,pad=0.3', facecolor='white', alpha=0.7)
        )

        plt.legend()
        plt.grid(True)
        #plt.show()        

    def show_plots(self) -> None:
        plt.show()

    # #implement these later whenever
    # def calculate_parameters_AxsquaredplusBxplusC(self, p: float=0, manual_delta_y: float=0.1) -> None:
    #     pass
    
    # def plot_AxsquaredplusBxplusC(self, p: float=0) -> None:
    #     pass
   
    # def calculate_parameters_Aoverxplusb(self, p: float=0, manual_delta_y: float=0.1) -> None:
    #     pass

    #     # #fit y=A(1/x)
    #     # print("FITTING y = (A/x)+b(?):")
    #     # number_of_parameters = 2
    #     # degrees_of_freedom = len(x_values) - number_of_parameters

    #     # global A
    #     # global delta_A
    #     # A = 
    #     # delta_A = 
        
    #     # chi_squared = sigma_sum(0, len(x_values)-1, lambda i: ((1/(delta_y_values[i]))*((y_values[i] - final_a_yaxplusb * x_values[i] - b)))**2)
    #     # reduced_chi_squared = chi_squared / degrees_of_freedom

    #     #Manually try better delta_y_values to get reduced_chi_squared of 0-1 (depending of degrees of freedom)
    #     # for i in range(len(delta_y_values)):
    #     #     delta_y_values[i] = 1.39

    #     # print(f"{delta_y_values=}, {delta_x_values=}, {A=}, {delta_A=}, \n {chi_squared=}, {reduced_chi_squared=}")
    
    # def plot_Aoverxplusb(self, p: float=0) -> None:
    #     pass

    # def calculate_parameters_Aexp_bxplusc_plusD(self, p: float=0, manual_delta_y: float=0.1) -> None:
    #     pass

    # def plot_Aexp_bxplusc_plusD(self, p: float=0) -> None:
        pass

def main() -> None:
    data_fitter_plotter = DataFitterPlotter('sampledata.xlsx', 'Sheet2')

    data_fitter_plotter.calculate_parameters_yax()
    data_fitter_plotter.plot_y_ax()

    data_fitter_plotter.calculate_parameters_yaxplusb()
    data_fitter_plotter.plot_y_axplusb()

    #Need reduced_chi_squared = 0.9xx for P=0.50 for ___ degrees of freedom
    data_fitter_plotter.calculate_parameters_yax(p=0.5, manual_delta_y=0.4)
    data_fitter_plotter.plot_y_ax(p=0.5)

    #Need reduced_chi_squared = 0.9xx for P=0.50 for ___ degrees of freedome
    data_fitter_plotter.calculate_parameters_yaxplusb(p=0.5, manual_delta_y=0.45)
    data_fitter_plotter.plot_y_axplusb(p=0.5)

    data_fitter_plotter.show_plots()

    # data_fitter_plotter.calculate_parameters_yaxSQUAREDplusb()
    #data_fitter_plotter.plot_y_axSQURAEDSSSDSDDplusb()

    # data_fitter_plotter.calculate_parameters_yaxSQUAREDplusb(p...)
    #data_fitter_plotter.plot_y_axSQURAEDSSSDSDDplusb(p....)


    # When implemented (optional):
    # calculate_parameters_Axsquaredplusbxplusc(0)
    # plot...
    #
    # calculate_parameters_Aoverxplusb(0)
    # plot...
    #
    # calculate_parameters_Aexp_bxplusc_plusD(0)
    # plot...

if __name__ == '__main__':
    main()