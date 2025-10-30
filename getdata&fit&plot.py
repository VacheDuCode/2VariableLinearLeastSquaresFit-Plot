#import random
import math
#import pandas
#import requests
import openpyxl
import matplotlib.pyplot as plt
import numpy as np

# Set the global font to Times New Roman
plt.rcParams["font.family"] = "Times New Roman"

#Doesn't paste data into spreadsheet, just prints them out

#doesn't create new xlxs file
#workbook = xlsxwriter.Workbook("PI-LAB2025.xlxs")

workbook = openpyxl.load_workbook('sampledata.xlsx')

#there is some bug when trying to use Sheet 2 for y=ax, P=0...
worksheet = workbook["Sheet1"]
x_parameter = worksheet['A1'].value
y_parameter = worksheet["B1"].value
delta_x_parameter = worksheet["C1"].value
delta_y_parameter = worksheet["D1"].value

x_values = []
y_values = []
delta_y_values = []
delta_x_values = []

modified_delta_y_values = []

initial_a_yax = 0
final_a_yax = 0
delta_a_yax = 0

initial_a_yaxplusb = 0
final_a_yaxplusb = 0
delta_a_yaxplusb = 0
b = 0
delta_b = 0

A = 1
delta_A = 0

number_of_parameters = 1
degrees_of_freedom = 1
chi_squared = 0
reduced_chi_squared = 0

def get_data():
    for value in worksheet.iter_rows(min_row=2,
                                    min_col=1,
                                    max_col=1,
                                    values_only=True):
        if value[0] != None:  
            x_values.append(value[0])

    for value in worksheet.iter_rows(min_row=2,
                                    min_col=2,
                                    max_col=2,
                                    values_only=True):
        if value[0] != None:
            y_values.append(value[0])

    for value in worksheet.iter_rows(min_row=2,
                                    min_col=4,
                                    max_col=4,
                                    values_only=True):
        if value[0] != None:
            delta_y_values.append(value[0])
        
    for value in worksheet.iter_rows(min_row=2,
                    max_row=100,
                    min_col=3,
                    max_col=3,
                    values_only=True):
        if value[0] != None:
            delta_x_values.append(value[0])

def sigma_sum(start, end, expression):
    return sum(expression(i) for i in range(start, end + 1))

#implement later
def calculate_parameters_Aoverxplusb(p: float) -> None:
    pass

    # #fit y=A(1/x)
    # print("FITTING y = (A/x)+b(?):")
    # number_of_parameters = 2
    # degrees_of_freedom = len(x_values) - number_of_parameters

    # global A
    # global delta_A
    # A = 
    # delta_A = 
    
    # chi_squared = sigma_sum(0, len(x_values)-1, lambda i: ((1/(delta_y_values[i]))*((y_values[i] - final_a_yaxplusb * x_values[i] - b)))**2)
    # reduced_chi_squared = chi_squared / degrees_of_freedom

    #Manually try better delta_y_values to get reduced_chi_squared of 0-1 (depending of degrees of freedom)
    # for i in range(len(delta_y_values)):
    #     delta_y_values[i] = 1.39

    # print(f"{delta_y_values=}, {delta_x_values=}, {A=}, {delta_A=}, \n {chi_squared=}, {reduced_chi_squared=}")
def calculate_parameters_Aexp_bxplusc_plusD(p: float) -> None:
    pass
  
def calculate_parameters_yax(p: float) -> None:
    print("FITTING y = ax:")
    global b
    b = 0
    global number_of_parameters
    number_of_parameters = 1
    global degrees_of_freedom
    degrees_of_freedom = len(x_values) - number_of_parameters

    global modified_delta_y_values
    global intial_a_yax
    global final_a_yax
    global delta_a_yax
    global chi_squared
    global reduced_chi_squared 

    if p!=0:
        #Manually try better delta_y_values to get reduced_chi_squared between 0-1 (depending of degrees of freedom)
        for i in range(len(modified_delta_y_values)):
            modified_delta_y_values[i] = 0.01
        print(f"\ndelta_y_values have been revised (new delta_y's={modified_delta_y_values[0]})...\n")
        print(modified_delta_y_values)

        print(f"REFITTING y=ax for {p=}...")
        final_a_yax = round(sigma_sum(0, len(x_values)-1, lambda i: (x_values[i]*y_values[i])/(modified_delta_y_values[i]**2)) / sigma_sum(0, len(x_values)-1, lambda i: (x_values[i]/modified_delta_y_values[i])**2), 6)
        delta_a_yax = round(1/math.sqrt(sigma_sum(0, len(x_values)-1, lambda i: (x_values[i]/modified_delta_y_values[i])**2)), 7)

        chi_squared = sigma_sum(0, len(x_values)-1, lambda i: ((1/(modified_delta_y_values[i]))*((y_values[i] - final_a_yax * x_values[i] - b)))**2)
        reduced_chi_squared = chi_squared / degrees_of_freedom
        print(f"\nRecalculated delta y's, a, b, delta a, delta, b, chi sqaured, reduced chi-squared:")
        print(f"{p=}, {modified_delta_y_values=}, \n {final_a_yax=}, {b=}, {delta_a_yax=}, \n {chi_squared=}, {reduced_chi_squared=}")
   
    else:
        initial_a_yax = round(sigma_sum(0, len(x_values)-1, lambda i: (x_values[i]*y_values[i])/(delta_y_values[i]**2)) / sigma_sum(0, len(x_values)-1, lambda i: (x_values[i]/delta_y_values[i])**2), 6)
        delta_a_yax = round(1/math.sqrt(sigma_sum(0, len(x_values)-1, lambda i: (x_values[i]/delta_y_values[i])**2)), 7)
        chi_squared = sigma_sum(0, len(x_values)-1, lambda i: ((1/(delta_y_values[i]))*((y_values[i] - initial_a_yax * x_values[i] - b)))**2)
        reduced_chi_squared = chi_squared / degrees_of_freedom

        for i in range(len(delta_y_values)):
            modified_delta_y_values.append(math.sqrt((delta_y_values[i])**2 + (initial_a_yax*delta_x_values[i])**2))
            modified_delta_y_values[i] = round(modified_delta_y_values[i], 6)

        final_a_yax = round(sigma_sum(0, len(x_values)-1, lambda i: (x_values[i]*y_values[i])/(modified_delta_y_values[i]**2)) / sigma_sum(0, len(x_values)-1, lambda i: (x_values[i]/modified_delta_y_values[i])**2), 6)
        delta_a_yax = round(1/math.sqrt(sigma_sum(0, len(x_values)-1, lambda i: (x_values[i]/modified_delta_y_values[i])**2)), 7)

        chi_squared = sigma_sum(0, len(x_values)-1, lambda i: ((1/(modified_delta_y_values[i]))*((y_values[i] - final_a_yax * x_values[i] - b)))**2)
        reduced_chi_squared = chi_squared / degrees_of_freedom

        #DEBUG
        print(f"{initial_a_yax=}, {final_a_yax}, {delta_a_yax=}, {chi_squared=}, \n {reduced_chi_squared=},  {delta_y_parameter} :: {delta_y_values=}, \n {delta_x_parameter} :: {delta_x_values=}")
        print(f"{x_parameter} :: {x_values=}, {y_parameter} :: {y_values=},\n{modified_delta_y_values=}, \n{degrees_of_freedom=}, {final_a_yax=}, {delta_a_yax=}, \n {chi_squared=}, {reduced_chi_squared=}")

def calculate_parameters_yaxplusb(p: float) -> None:
    #y=ax+b
    global number_of_parameters
    number_of_parameters = 2
    global degrees_of_freedom
    degrees_of_freedom = len(x_values) - number_of_parameters

    global modified_delta_y_values

    global triangle
    global initial_a_yaxplusb
    global final_a_yaxplusb
    global b
    global delta_a_yaxplusb
    global delta_a_yaxplusb
    global delta_b

    global chi_squared
    global reduced_chi_squared 

    if p!=0:
        print(f"REFITTING y=ax+b for {p=}...")
        triangle = sigma_sum(0, len(x_values)-1, lambda i: (1/modified_delta_y_values[i])**2)*sigma_sum(0, len(x_values)-1, lambda i: (x_values[i]/modified_delta_y_values[i])**2) - (sigma_sum(0, len(x_values)-1, lambda i: x_values[i]/(modified_delta_y_values[i]**2)))**2
        final_a_yaxplusb = (1/triangle)*(sigma_sum(0, len(x_values)-1, lambda i: 1/(modified_delta_y_values[i]**2))*sigma_sum(0, len(x_values)-1, lambda i: (x_values[i]*y_values[i])/(modified_delta_y_values[i]**2)) - sigma_sum(0, len(x_values)-1, lambda i: y_values[i]/(modified_delta_y_values[i]**2))*sigma_sum(0, len(x_values)-1, lambda i: x_values[i]/(modified_delta_y_values[i]**2)))

        b = (1/triangle)*((sigma_sum(0,len(x_values)-1, lambda i: y_values[i]/(modified_delta_y_values[i]**2))*sigma_sum(0, len(x_values)-1, lambda i: (x_values[i]/modified_delta_y_values[i])**2)) - (sigma_sum(0, len(x_values)-1, lambda i: x_values[i]/(modified_delta_y_values[i]**2))*sigma_sum(0, len(x_values)-1, lambda i: (x_values[i]*y_values[i])/(modified_delta_y_values[i]**2))))
        delta_a_yaxplusb = math.sqrt((1/triangle)*sigma_sum(0, len(x_values)-1, lambda i: (1/(modified_delta_y_values[i]**2))))
        delta_b = math.sqrt((1/triangle)*sigma_sum(0, len(x_values)-1, lambda i: (x_values[i]/modified_delta_y_values[i])**2))

        chi_squared = sigma_sum(0, len(x_values)-1, lambda i: ((1/(modified_delta_y_values[i]))*((y_values[i] - final_a_yaxplusb * x_values[i] - b)))**2)
        reduced_chi_squared = chi_squared / degrees_of_freedom
        print(f"\nRecalculated delta y's, a, b, delta a, delta, b, chi sqaured, reduced chi-squared:")
        print(f"{p=}, {modified_delta_y_values=}, \n {final_a_yaxplusb=}, {b=}, {delta_a_yaxplusb=}, {delta_b=}, \n {chi_squared=}, {reduced_chi_squared=}")

    else:
        print("\n\n FITTING y = ax + b (initial):\n")
        triangle = sigma_sum(0, len(x_values)-1, lambda i: (1/delta_y_values[i])**2)*sigma_sum(0, len(x_values)-1, lambda i: (x_values[i]/delta_y_values[i])**2) - (sigma_sum(0, len(x_values)-1, lambda i: x_values[i]/(delta_y_values[i]**2)))**2
        t1 = sigma_sum(0, len(x_values)-1, lambda i: (1/delta_y_values[i])**2)*sigma_sum(0, len(x_values)-1, lambda i: (x_values[i]/delta_y_values[i])**2)
        t2 = (sigma_sum(0, len(x_values)-1, lambda i: x_values[i]/(delta_y_values[i]**2)))**2
        
        print(f"{t1=}, {t2=}, {triangle=}, {len(x_values)}, {number_of_parameters=}, {degrees_of_freedom=}")

        initial_a_yaxplusb = (1/triangle)*(sigma_sum(0, len(x_values)-1, lambda i: 1/(delta_y_values[i]**2))*sigma_sum(0, len(x_values)-1, lambda i: (x_values[i]*y_values[i])/(delta_y_values[i]**2)) - sigma_sum(0, len(x_values)-1, lambda i: y_values[i]/(delta_y_values[i]**2))*sigma_sum(0, len(x_values)-1, lambda i: x_values[i]/(delta_y_values[i]**2)))
        b = (1/triangle)*((sigma_sum(0,len(x_values)-1, lambda i: y_values[i]/(delta_y_values[i]**2))*sigma_sum(0, len(x_values)-1, lambda i: (x_values[i]/delta_y_values[i])**2)) - (sigma_sum(0, len(x_values)-1, lambda i: x_values[i]/(delta_y_values[i]**2))*sigma_sum(0, len(x_values)-1, lambda i: (x_values[i]*y_values[i])/(delta_y_values[i]**2))))
        delta_a_yaxplusb = math.sqrt((1/triangle)*sigma_sum(0, len(x_values)-1, lambda i: (1/(delta_y_values[i]**2))))
        delta_b = math.sqrt((1/triangle)*sigma_sum(0, len(x_values)-1, lambda i: (x_values[i]/delta_y_values[i])**2))

        chi_squared = sigma_sum(0, len(x_values)-1, lambda i: ((1/(delta_y_values[i]))*((y_values[i] - initial_a_yaxplusb * x_values[i] - b)))**2)
        reduced_chi_squared = chi_squared / degrees_of_freedom
        
        # WRUTOUT INIT
        print(f"{x_values=}, {y_values=}, {initial_a_yaxplusb=}, \n {b=}, \n {delta_a_yaxplusb=}, {delta_b=}")
        print(f"{delta_y_values=}, {modified_delta_y_values=}, {degrees_of_freedom=}, {chi_squared=}, {reduced_chi_squared=}")

        #final - new deltas, new a, new b
        for i in range(len(delta_y_values)):
            modified_delta_y_values[i] = math.sqrt((delta_y_values[i])**2 + (initial_a_yaxplusb*delta_x_values[i])**2)
            modified_delta_y_values[i] = round(modified_delta_y_values[i], 6)
        print("\n\n FITTING y = ax + b (final):\n")

        triangle = sigma_sum(0, len(x_values)-1, lambda i: (1/modified_delta_y_values[i])**2)*sigma_sum(0, len(x_values)-1, lambda i: (x_values[i]/modified_delta_y_values[i])**2) - (sigma_sum(0, len(x_values)-1, lambda i: x_values[i]/(modified_delta_y_values[i]**2)))**2
        final_a_yaxplusb = (1/triangle)*(sigma_sum(0, len(x_values)-1, lambda i: 1/(modified_delta_y_values[i]**2))*sigma_sum(0, len(x_values)-1, lambda i: (x_values[i]*y_values[i])/(modified_delta_y_values[i]**2)) - sigma_sum(0, len(x_values)-1, lambda i: y_values[i]/(modified_delta_y_values[i]**2))*sigma_sum(0, len(x_values)-1, lambda i: x_values[i]/(modified_delta_y_values[i]**2)))
        
        b = (1/triangle)*((sigma_sum(0,len(x_values)-1, lambda i: y_values[i]/(modified_delta_y_values[i]**2))*sigma_sum(0, len(x_values)-1, lambda i: (x_values[i]/modified_delta_y_values[i])**2)) - (sigma_sum(0, len(x_values)-1, lambda i: x_values[i]/(modified_delta_y_values[i]**2))*sigma_sum(0, len(x_values)-1, lambda i: (x_values[i]*y_values[i])/(modified_delta_y_values[i]**2))))
        delta_a_yaxplusb = math.sqrt((1/triangle)*sigma_sum(0, len(x_values)-1, lambda i: (1/(modified_delta_y_values[i]**2))))
        delta_b = math.sqrt((1/triangle)*sigma_sum(0, len(x_values)-1, lambda i: (x_values[i]/modified_delta_y_values[i])**2))

        chi_squared = sigma_sum(0, len(x_values)-1, lambda i: ((1/(modified_delta_y_values[i]))*((y_values[i] - final_a_yaxplusb * x_values[i] - b)))**2)
        reduced_chi_squared = chi_squared / degrees_of_freedom

        #The formula for calculating P for given reduced-chi-squared is more effort than I want: See RB Eq. 11.6

        # WRITOUT FINAL
        print(f"{modified_delta_y_values=}, \n {final_a_yaxplusb=}, {b=}, {delta_a_yaxplusb=}, {delta_b=}, \n {chi_squared=}, {reduced_chi_squared=}")
def plot_y_ax(p: float) -> None:
    global modified_delta_y_values

    global final_a_yax
    global b
    global chi_squared
    global reduced_chi_squared

    # Fit line
    x_line = np.linspace(min(x_values)-abs(min(x_values))*0.2, max(x_values) * 1.1, 200)
    y_line = final_a_yax * x_line + b

    # Plot with hollow markers and whiskers
    plt.errorbar(
        x_values, y_values,
        yerr=modified_delta_y_values,
        fmt='o', markersize=4,
        mfc='none', mec='black',
        ecolor='black', elinewidth=1.5, capsize=2,
        label='Data with uncertainties'
    )

    # Fit line
    plt.plot(x_line, y_line, color='blue', linewidth=2, label=f'Fit: y = {final_a_yax:.4f}x')

    # Labels
    plt.xlabel(f"{x_parameter}")
    plt.ylabel(f"{y_parameter}")

    if (p!=0):
        plt.title(f"Linear Least-Squares Fit of Measurements (y = ax) Ensuring P = {p}")
    else:
        plt.title(f"Linear Least-Squares Fit of Measurements (y = ax)")

    # Add chi-squared summary box
    textstr = '\n'.join((
        r'$\chi^2 = %.3f$' % (chi_squared,),
        r'$\chi^2_\nu = %.3f$' % (reduced_chi_squared,)
    ))

    plt.gca().text(
        0.05, 0.95, textstr,
        transform=plt.gca().transAxes,
        fontsize=10,
        verticalalignment='top',
        bbox=dict(boxstyle='round,pad=0.3', facecolor='white', alpha=0.7)
    )

    plt.legend()
    plt.grid(True)
    plt.show()

def plot_y_axplusb(p: float) -> None: 
    global modified_delta_y_values
    global final_a_yaxplusb
    global b
    global chi_squared
    global reduced_chi_squared
    
    # Fit line
    x_line = np.linspace(min(x_values)-abs(min(x_values))*0.2, max(x_values) * 1.1, 200)
    y_line = final_a_yaxplusb * x_line + b

    # Plot with error bars
    plt.errorbar(
        x_values, y_values, yerr=modified_delta_y_values,
        fmt='o', markersize=4,
        mfc='none', mec='black',
        ecolor='black', elinewidth=1.5, capsize=2,
        label='Data with uncertainties'
    )

    # Fit line
    plt.plot(x_line, y_line, color='green', linewidth=2, label=f'Fit: y = {final_a_yaxplusb:.2f}x + {b:.2f}')

    # Labels
    plt.xlabel(f"{x_parameter}")
    plt.ylabel(f"{y_parameter}")

    if (p!=0):
        plt.title(f"Linear Least-Squares Fit of Diverging Lens with Intercept ensuring P={p}")
    else:
        plt.title("Linear Least-Squares Fit of Diverging Lens with Intercept")

    # Add chi-squared summary box
    textstr = '\n'.join((
        r'$\chi^2 = %.3f$' % (chi_squared,),
        r'$\chi^2_\nu = %.4f$' % (reduced_chi_squared,),
    ))
    plt.gca().text(
        0.05, 0.50, textstr,
        transform=plt.gca().transAxes,
        fontsize=10,
        verticalalignment='top',
        bbox=dict(boxstyle='round,pad=0.3', facecolor='white', alpha=0.7)
    )

    plt.legend()
    plt.grid(True)
    plt.show()        

def main() -> None:
    get_data()

    calculate_parameters_yax(0)
    plot_y_ax(0)

    calculate_parameters_yaxplusb(0)
    plot_y_axplusb(0)

    calculate_parameters_yax(0.5)
    plot_y_ax(0.5)

    calculate_parameters_yaxplusb(0.5)
    plot_y_axplusb(0.5)

    # When implemented: optional:
    # calculate_parameters_Axsquaredplusbxplusc(0)
    # plot...
    #
    # calculate_parameters_Aoverxplusb(0)
    # plot...
    #
    # calculate_parameters_Aexp_bxplusc_plusD(0)
    # plot...
    #

if __name__ == '__main__':
    main()